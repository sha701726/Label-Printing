# Individual packaging label

from PIL import Image, ImageDraw, ImageFont
import barcode
from barcode.writer import ImageWriter
from openpyxl import load_workbook
from datetime import datetime

# -------------------------------
# Load Excel
# -------------------------------
wb = load_workbook("ipl.xlsx")
sheet = wb.active

# -------------------------------
# Font (BOLD)
# -------------------------------
try:
    font = ImageFont.truetype("C:/Windows/Fonts/arialbd.ttf", 40)
    font_small = ImageFont.truetype("C:/Windows/Fonts/arialbd.ttf", 36)
except:
    font = ImageFont.load_default()
    font_small = font

# -------------------------------
# ROI BASE
# Outer border: (1057, 914, 1445, 1221)
# -------------------------------
BASE_X = 1057
BASE_Y = 914
WIDTH = 1445
HEIGHT = 1221

# -------------------------------
# Barcode function (NO TEXT)
# -------------------------------
def generate_barcode(data, size):
    if not data:
        return Image.new("RGB", size, "white")

    code = barcode.get('code128', str(data), writer=ImageWriter())
    img = code.render(writer_options={
        'write_text': False,
        'module_width': 0.35,
        'module_height': 20
    })
    return img.resize(size)

# -------------------------------
# Date formatter (MM/DD/YYYY)
# -------------------------------
def format_date(value):
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").strftime("%m/%d/%Y")
    except:
        return str(value)

# -------------------------------
# Loop Excel Rows
# -------------------------------
for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):

    lam_pn = str(row[0] or "")
    sn = str(row[1] or "")
    mfg_date = format_date(row[2])
    rev = str(row[3] or "")
    coo = str(row[4] or "")
    qty = str(row[5] or "")

    # -------------------------------
    # Create canvas
    # -------------------------------
    img = Image.new("RGB", (WIDTH, HEIGHT), "white")
    draw = ImageDraw.Draw(img)

    def adj(x, y):
        return (x - BASE_X, y - BASE_Y)

    # -------------------------------
    # (1) Outer Border
    # -------------------------------
    draw.rectangle((0, 0, WIDTH-1, HEIGHT-1), outline="black", width=4)

    # -------------------------------
    # (2) LAM P/N TEXT
    # -------------------------------
    draw.text(adj(1095, 979), f"LAM P/N: {lam_pn}", fill="black", font=font)

    # -------------------------------
    # (3) LAM P/N BARCODE
    # -------------------------------
    img.paste(generate_barcode(lam_pn, (942, 250)), adj(1104, 1075))

    # -------------------------------
    # (4) S/N TEXT
    # -------------------------------
    draw.text(adj(1108, 1386), f"S/N: {sn}", fill="black", font=font)

    # -------------------------------
    # (5) S/N BARCODE
    # -------------------------------
    img.paste(generate_barcode(sn, (942, 192)), adj(1100, 1482))

    # -------------------------------
    # (6) MFG DATE
    # -------------------------------
    draw.text(adj(1385, 1901), f"MFG DATE: {mfg_date}", fill="black", font=font_small)

    # -------------------------------
    # (7) LOGO
    # -------------------------------
    try:
        logo = Image.open("logo.png").convert("RGBA")
        logo = logo.resize((328, 134))
        img.paste(logo, adj(2131, 972), logo)
    except:
        pass

    # -------------------------------
    # (8) REV
    # -------------------------------
    draw.text(adj(2259, 1252), f"REV: {rev}", fill="black", font=font_small)

    # -------------------------------
    # (9) CoO
    # -------------------------------
    draw.text(adj(2216, 1409), f"CoO: {coo}", fill="black", font=font_small)

    # -------------------------------
    # (10) QTY
    # -------------------------------
    draw.text(adj(2246, 1647), f"QTY: {qty}", fill="black", font=font_small)

    # -------------------------------
    # SAVE
    # -------------------------------
    img.save(f"logo{lam_pn}_{i}.png")

print("Logo labels generated successfully!")