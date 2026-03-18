#Outer Packege  with label

from PIL import Image, ImageDraw, ImageFont
import barcode
from barcode.writer import ImageWriter
from openpyxl import load_workbook

# -------------------------------
# Load Excel
# -------------------------------
wb = load_workbook("opwl.xlsx")
sheet = wb.active

# -------------------------------
# Font (BOLD + smaller size)
# -------------------------------
try:
    font = ImageFont.truetype("C:/Windows/Fonts/arialbd.ttf", 32)
    font_big = ImageFont.truetype("C:/Windows/Fonts/arialbd.ttf", 38)
except:
    font = ImageFont.load_default()
    font_big = font

# -------------------------------
# ROI BASE
# -------------------------------
BASE_X = 1286
BASE_Y = 829
WIDTH = 1420
HEIGHT = 1498

# -------------------------------
# Barcode function (NO TEXT BELOW)
# -------------------------------
def generate_barcode(data, size):
    if not data:
        return Image.new("RGB", size, "white")

    code = barcode.get('code128', str(data), writer=ImageWriter())
    img = code.render(writer_options={
        'write_text': False,
        'module_height': 15,
        'module_width': 0.35,
        'quiet_zone': 2
    })
    return img.resize(size)

# -------------------------------
# Loop Excel Rows
# -------------------------------
for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):

    # -------------------------------
    # SAFE DATA EXTRACTION
    # -------------------------------
    lam_pn = str(row[1] or "")
    sn = str(row[2] or "")
    po = str(row[3] or "")
    ship_date = str(row[4] or "")
    container = str(row[5] or "")

    # FIXED COLUMN MAPPING (G, H, I, J)
    rev = str(row[6] or "") if len(row) > 6 else ""
    coo = str(row[7] or "") if len(row) > 7 else ""
    qty = str(row[8] or "") if len(row) > 8 else ""
    po_line = str(row[9] or "") if len(row) > 9 else ""

    # -------------------------------
    # Create label
    # -------------------------------
    img = Image.new("RGB", (WIDTH, HEIGHT), "white")
    draw = ImageDraw.Draw(img)

    def adj(x, y):
        return (x - BASE_X, y - BASE_Y)

    # -------------------------------
    # TEXT SECTION
    # -------------------------------
    draw.text(adj(1351, 899), "SHIP TO:", fill="black", font=font_big)

    # Address Box
    x, y = adj(1351, 960)
    draw.rectangle((x, y, x + 894, y + 315), outline="black", width=3)

    draw.text(adj(1355, 1332), f"LAM P/N: {lam_pn}", fill="black", font=font)
    draw.text(adj(1359, 1586), f"S/N: {sn}", fill="black", font=font)
    draw.text(adj(1363, 1843), f"PURCHASE ORDER: {po}", fill="black", font=font)

    draw.text(adj(1359, 2081), f"SHIP DATE: {ship_date}", fill="black", font=font)
    draw.text(adj(1355, 2162), "PACKING LIST ENCLOSED", fill="black", font=font)
    draw.text(adj(1359, 2231), f"CONTAINER: {container}", fill="black", font=font)

    # -------------------------------
    # EXTRA FIELDS (RIGHT SIDE)
    # -------------------------------
    draw.text(adj(2428, 1490), f"REV: {rev}", fill="black", font=font)
    draw.text(adj(2397, 1590), f"CoO: {coo}", fill="black", font=font)
    draw.text(adj(2424, 1736), f"QTY: {qty}", fill="black", font=font)
    draw.text(adj(2180, 1985), f"PO LINE ITEM: {po_line}", fill="black", font=font)

    # -------------------------------
    # BARCODES
    # -------------------------------
    img.paste(generate_barcode(lam_pn, (1019, 139)), adj(1359, 1405))
    img.paste(generate_barcode(sn, (1019, 154)), adj(1359, 1651))
    img.paste(generate_barcode(po, (802, 119)), adj(1370, 1901))

    # -------------------------------
    # LOGO (optional)
    # -------------------------------
    try:
        logo = Image.open("logo.png").convert("RGBA")
        logo = logo.resize((200, 80))
        img.paste(logo, adj(2317, 883), logo)
    except:
        pass

    # -------------------------------
    # SAVE
    # -------------------------------
    img.save(f"label_{i}.png")

print(" All labels generated successfully!")